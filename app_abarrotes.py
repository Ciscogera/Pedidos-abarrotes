import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import json
import os
import urllib.parse
from datetime import datetime

# --- CONFIGURACIÓN DE PÁGINA WEB ---
st.set_page_config(page_title="Pedidos Cocina - El Bajo", page_icon=":material/local_pizza:", layout="centered")

# --- CATÁLOGO DE PRODUCTOS INTEGRADO ---
PRODUCTOS_BASE = [
    {"Nombre": "Limon en malla", "Par": 8.0, "Medida": "MALLAS"},
    {"Nombre": "Pomelo", "Par": 10.0, "Medida": "KG"},
    {"Nombre": "Naranja", "Par": 15.0, "Medida": "KG"},
    {"Nombre": "Limon Sutil", "Par": 8.0, "Medida": "KG"},
    {"Nombre": "Frutillas", "Par": 4.0, "Medida": "KG"},
    {"Nombre": "piña", "Par": 2.0, "Medida": "UND"},
    {"Nombre": "Jengibre", "Par": 2.0, "Medida": "KG"},
    {"Nombre": "Menta", "Par": 2.0, "Medida": "KG"},
    {"Nombre": "Pera", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Pepino", "Par": 3.0, "Medida": "UND"},
    {"Nombre": "Apio", "Par": 1.0, "Medida": "UND"},
    {"Nombre": "Flores comestibles", "Par": 6.0, "Medida": "UND"},
    {"Nombre": "Aceitunas verdes sin carozo", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Ramas Canela", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Flor de Jamaica(Hibisco deshidatrado)", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Hibisco confitao", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Rosa rugosa", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Chancaca", "Par": 2.0, "Medida": "KG"},
    {"Nombre": "Huesillo", "Par": 3.0, "Medida": "KG"},
    {"Nombre": "Azucar Blanca", "Par": 30.0, "Medida": "KG"},
    {"Nombre": "Azucar rubia", "Par": 3.0, "Medida": "KG"},
    {"Nombre": "Maraschinno", "Par": 10.0, "Medida": "UND"},
    {"Nombre": "Gysophilia", "Par": 6.0, "Medida": "UND"},
    {"Nombre": "Jugo de Tomate", "Par": 10.0, "Medida": "UND"},
    {"Nombre": "Leche Sin Lactosa", "Par": 12.0, "Medida": "LT"},
    {"Nombre": "Crema de Leche", "Par": 6.0, "Medida": "LT"},
    {"Nombre": "Cola cao", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Endulzante", "Par": 6.0, "Medida": "UND"},
    {"Nombre": "Tabasco", "Par": 12.0, "Medida": "UND"},
    {"Nombre": "Sal", "Par": 6.0, "Medida": "KG"},
    {"Nombre": "Merken", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Anís estrella", "Par": 0.5, "Medida": "KG"},
    {"Nombre": "Salsa Inglesa", "Par": 12.0, "Medida": "UND"},
    {"Nombre": "Papa Lays", "Par": 8.0, "Medida": "UND"},
    {"Nombre": "Coco Lopez", "Par": 6.0, "Medida": "UND"},
    {"Nombre": "Jugo en sobre", "Par": 60.0, "Medida": "UND"},
    {"Nombre": "Nescafe", "Par": 4.0, "Medida": "UND"},
    {"Nombre": "Té/Hierbas", "Par": 1.0, "Medida": "CAJA"},
    {"Nombre": "Aji puta madre", "Par": 0.2, "Medida": "KG"},
    {"Nombre": "Arandanos", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Albahaca", "Par": 1.0, "Medida": "KG"},
    {"Nombre": "Butterfly Pea", "Par": 0.4, "Medida": "KG"}
]

# --- CONTROL DE SESIÓN GENERAL ---
if "telefono_proveedor" not in st.session_state:
    st.session_state.telefono_proveedor = "569"
if "responsable" not in st.session_state:
    st.session_state.responsable = ""
if "fecha_inventario" not in st.session_state:
    st.session_state.fecha_inventario = datetime.today().strftime('%Y-%m-%d')
if "paso_flujo" not in st.session_state:
    st.session_state.paso_flujo = "configuracion"
if "datos_borrador" not in st.session_state:
    st.session_state.datos_borrador = {}
if "excel_final" not in st.session_state:
    st.session_state.excel_final = None

# --- FUNCIONES DE PERSISTENCIA ANTI-CAÍDAS ---
def obtener_nombre_archivo_borrador(responsable, fecha):
    nombre_limpio = responsable.lower().strip().replace(" ", "_")
    return f"borrador_{nombre_limpio}_{fecha}.json"

def guardar_borrador_en_disco(responsable, fecha):
    """Escanea el estado de los widgets y escribe un respaldo de seguridad en tiempo real."""
    archivo = obtener_nombre_archivo_borrador(responsable, fecha)
    datos_progreso = {}
    for key in st.session_state.keys():
        if key.startswith("stock_"):
            clave_producto = key.replace("stock_", "")
            datos_progreso[clave_producto] = st.session_state[key]
            
    with open(archivo, "w", encoding="utf-8") as f:
        json.dump(datos_progreso, f, ensure_ascii=False, indent=4)

def cargar_borrador_de_disco(responsable, fecha):
    """Busca si el usuario tenía un conteo a medias antes de que se cortara el internet."""
    archivo = obtener_nombre_archivo_borrador(responsable, fecha)
    if os.path.exists(archivo):
        with open(archivo, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

# --- CLASIFICADOR DINÁMICO ---
def categorizar_producto(nombre):
    nombre_lower = nombre.lower()
    verduleria = ["limon", "pomelo", "naranja", "frutilla", "piña", "jengibre", "menta", "pera", "pepino", "apio", "flores comestibles", "aji", "arandanos", "albahaca"]
    especias = ["canela", "jamaica", "hibisco", "rosa rugosa", "chancaca", "huesillo", "merken", "anís", "butterfly"]
    if any(x in nombre_lower for x in verduleria): return "Frutas y Verduras Frescas"
    elif any(x in nombre_lower for x in especias): return "Especias y Deshidratados"
    else: return "Abarrotes, Lácteos y Otros"

# --- CONSTRUCTOR DINÁMICO DE EXCEL ---
def generar_excel_desde_datos(conteos, responsable, fecha):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    font_titulo = Font(name="Calibri", size=14, bold=True, color="1B365D")
    font_sub = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    border_thin = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    
    ws["A1"] = "Inventario Abarrote Barra"
    ws["A1"].font = font_titulo
    ws["A2"] = f"Nombre Responsable: {responsable}"
    ws["A2"].font = font_sub
    ws["A3"] = f"Fecha: {fecha}"
    ws["A3"].font = font_sub
    
    headers = ["Nombre", "Par Stock", "Bodega", "MEDIDA", "Pedido", "MEDIDA"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, (clave, info) in enumerate(conteos.items(), 5):
        actual = info["Actual"]
        cantidad_pedir = max(0.0, info["Par"] - actual)
        
        ws.cell(row=row_idx, column=1, value=info["Producto"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2, value=info["Par"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=3, value=actual if actual > 0 else "").alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=4, value=info["Medida"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=5, value=cantidad_pedir if cantidad_pedir > 0 else "").alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=6, value=info["Medida"]).alignment = Alignment(horizontal="center")
        
        for col_idx in range(1, 7): ws.cell(row=row_idx, column=col_idx).border = border_thin
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# --- REDACTOR DE MENSAJE WHATSAPP ---
def generar_mensaje_whatsapp(lista_pedidos):
    mensaje = "*Hola, para la barra pediremos:*\n\n"
    categorias_presentes = lista_pedidos["Categoria"].unique()
    
    for cat in categorias_presentes:
        items_cat = lista_pedidos[lista_pedidos["Categoria"] == cat]
        mensaje += f"📦 *{cat.upper()}*\n"
        for _, row in items_cat.iterrows():
            cant = row['Cantidad_Pedir']
            cant_str = f"{int(cant)}" if cant % 1 == 0 else f"{cant}"
            mensaje += f"• {row['Producto']}: *{cant_str}* {row['Medida']}\n"
        mensaje += "\n"
        
    mensaje += "Muchas gracias!"
    return mensaje

# --- INTERFAZ ---
st.title("Pedidos de Abarrotes - El Bajo")

# =====================================================================
# PASO 1: CONFIGURACIÓN OBLIGATORIA
# =====================================================================
if st.session_state.paso_flujo == "configuracion":
    st.subheader(":material/vpn_key: Configuración Obligatoria de Turno")
    
    responsable_temp = st.text_input("📝 Tu Nombre (Responsable del turno):", value=st.session_state.responsable, placeholder="Ej: Carlos Silva")
    telefono_temp = st.text_input("📞 WhatsApp del Proveedor (Celular de 9 dígitos o con +569):", value=st.session_state.telefono_proveedor, placeholder="Ej: 9 1234 5678")
    fecha_temp = st.date_input("📅 Fecha del Inventario:", value=datetime.strptime(st.session_state.fecha_inventario, '%Y-%m-%d')).strftime('%Y-%m-%d')
    
    st.markdown("---")
    
    if st.button("INICIAR CONTEO DE STOCK", icon=":material/checklist:", use_container_width=True):
        tel_limpio = "".join(c for c in telefono_temp if c.isdigit())
        if len(tel_limpio) == 9 and tel_limpio.startswith("9"): tel_limpio = "56" + tel_limpio
        elif len(tel_limpio) == 8: tel_limpio = "569" + tel_limpio
            
        if not responsable_temp.strip():
            st.error("⚠️ Debes ingresar tu nombre.")
        elif len(tel_limpio) != 11 or not tel_limpio.startswith("569"):
            st.error("⚠️ Número de WhatsApp no válido.")
        else:
            st.session_state.responsable = responsable_temp.strip()
            st.session_state.telefono_proveedor = tel_limpio
            st.session_state.fecha_inventario = fecha_temp
            
            # 🔍 INTELIGENCIA: Intentamos cargar un borrador existente para este usuario/fecha
            st.session_state.datos_borrador = cargar_borrador_de_disco(st.session_state.responsable, st.session_state.fecha_inventario)
            
            st.session_state.paso_flujo = "formulario"
            st.rerun()

# =====================================================================
# PASO 2: INVENTARIO ACTIVO
# =====================================================================
elif st.session_state.paso_flujo == "formulario":
    st.success(f":material/lock_open: **Turno Activo:** {st.session_state.responsable} | 📅 {st.session_state.fecha_inventario}")
    
    # 🚨 AVISO DE BORRADOR DETECTADO
    if st.session_state.datos_borrador:
        st.info(":material/restore_page: **Borrador detectado y cargado.** Tu avance anterior ha sido restaurado.", icon=":material/info:")
        
    col_config, col_save = st.columns(2)
    with col_config:
        if st.button("Configuración / Turno", icon=":material/settings:", use_container_width=True):
            st.session_state.paso_flujo = "configuracion"
            st.rerun()
    with col_save:
        # 💾 EL BOTÓN DE SALVAGUARDA MANUAL
        if st.button("GUARDAR PROGRESO", icon=":material/save:", type="primary", use_container_width=True):
            guardar_borrador_en_disco(st.session_state.responsable, st.session_state.fecha_inventario)
            st.toast("✅ ¡Progreso guardado a salvo en el servidor!", icon="💾")
        
    st.markdown("---")
    
    df_inv = pd.DataFrame(PRODUCTOS_BASE)
    df_inv["Categoria"] = df_inv["Nombre"].apply(categorizar_producto)
    
    st.subheader(":material/fact_check: Conteo Actual de Insumos")
    
    conteos_usuario = {}
    categorias = ["Frutas y Verduras Frescas", "Especias y Deshidratados", "Abarrotes, Lácteos y Otros"]
    iconos_material = {
        "Frutas y Verduras Frescas": ":material/nutrition: Frutas y Verduras Frescas",
        "Especias y Deshidratados": ":material/eco: Especias y Deshidratados",
        "Abarrotes, Lácteos y Otros": ":material/kitchen: Abarrotes, Lácteos y Otros"
    }
    
    for cat in categorias:
        df_cat = df_inv[df_inv["Categoria"] == cat]
        if df_cat.empty: continue
        
        titulo_seccion = iconos_material.get(cat, cat)
        
        with st.expander(f"{titulo_seccion} ({len(df_cat)} productos)"):
            for idx, row in df_cat.iterrows():
                prod_name = row["Nombre"]
                par_stock = row["Par"]
                medida = row["Medida"]
                
                st.markdown(f"**{prod_name}**")
                st.caption(f"📏 Medida: {medida} | Par: {par_stock}")
                
                decimales_activos = (par_stock % 1 != 0)
                if decimales_activos:
                    step_val, min_val, val_format = 0.1, 0.0, "%.1f"
                else:
                    step_val, min_val, val_format = 1, 0, "%d"
                
                # 🔍 COMPROBACIÓN DE BORRADOR: Extraemos el valor previo si es que existe
                clave_prod = f"{idx}_{prod_name}"
                if clave_prod in st.session_state.datos_borrador:
                    val_inicial = float(st.session_state.datos_borrador[clave_prod]) if decimales_activos else int(st.session_state.datos_borrador[clave_prod])
                else:
                    val_inicial = min_val
                
                actual_val = st.number_input(
                    label=f"En Stock ({medida})",
                    min_value=min_val,
                    value=val_inicial, # Cargamos el valor rescatado o 0
                    step=step_val,
                    format=val_format,
                    key=f"stock_{idx}_{prod_name}"
                )
                
                conteos_usuario[clave_prod] = {
                    "Producto": prod_name,
                    "Par": par_stock,
                    "Actual": actual_val,
                    "Medida": medida,
                    "Categoria": cat
                }
                st.markdown("---")
                
    # --- PROCESAR PEDIDO FINAL ---
    if st.button("CALCULAR PEDIDO DE COCINA", icon=":material/calculate:", use_container_width=True):
        registros_pedido = []
        for clave, info in conteos_usuario.items():
            cantidad_pedir = max(0.0, info["Par"] - info["Actual"])
            if cantidad_pedir > 0:
                registros_pedido.append({
                    "Categoria": info["Categoria"],
                    "Producto": info["Producto"],
                    "Cantidad_Pedir": cantidad_pedir,
                    "Medida": info["Medida"]
                })
                
        df_pedido = pd.DataFrame(registros_pedido)
        
        if not df_pedido.empty:
            # Al calcular con éxito, también respaldamos por seguridad
            guardar_borrador_en_disco(st.session_state.responsable, st.session_state.fecha_inventario)
            
            st.session_state.excel_final = generar_excel_desde_datos(conteos_usuario, st.session_state.responsable, st.session_state.fecha_inventario)
            
            st.subheader(":material/list_alt: Pedido Sugerido")
            st.dataframe(df_pedido[["Categoria", "Producto", "Cantidad_Pedir", "Medida"]], use_container_width=True, hide_index=True)
            
            texto_pedido = generar_mensaje_whatsapp(df_pedido)
            st.subheader(":material/chat: Vista Previa del Pedido")
            st.code(texto_pedido, language="text")
            
            texto_codificado = urllib.parse.quote(texto_pedido)
            url_whatsapp = f"https://wa.me/{st.session_state.telefono_proveedor}?text={texto_codificado}"
            
            st.markdown("### 📲 Despacho Directo")
            st.link_button("ENVIAR PEDIDO POR WHATSAPP", url_whatsapp, icon=":material/send:", use_container_width=True)
            
            st.download_button(label="DESCARGAR PLANILLA EXCEL REGISTRADA", data=st.session_state.excel_final, file_name=f"Pedido_Cocina_{st.session_state.fecha_inventario}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        else:
            st.success("🎉 ¡Todos los niveles de stock están completos! No es necesario realizar pedidos hoy.")